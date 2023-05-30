# Author: Maksim Dmitriev : GitHub: Mak7bit
# Date: 21/02/2024 - For Osmose Australia Pty Ltd
# extractor_consequitive.csv - allows to read and filter Western Power pole data to produce a smaller data set containing onlly poles from selected contracts. 
#           Version: V.01 - Initial Version of code written following pseudo code. 
#                    V.02 - Opening a given folder and trying everything zip file or csv for the requested Contract Numbers
#                    V.03 - Does not filter thought Contract Numbers and creates a merged .xlsx file
#
#
#
#
#   Requirements to run: python extractor_consequitive_v03.py 








import sys
import zipfile
from openpyxl import load_workbook
import pandas as pd
import os

#def read_csv_from_zip(zip_file_name):
#    with zipfile.ZipFile(zip_file_name, 'r') as zip_ref:
#        files = zip_ref.namelist()
#        csv_file_name = None
#        for file in files:
#            if file.endswith('.csv'):
#                csv_file_name = file
#                break
#        if csv_file_name:
#            csv_file = zip_ref.open(csv_file_name)
#            df = pd.read_csv(csv_file)
#            return df
#        else:
#            return None
        
def read_csv_from_zip(file_path):
    with zipfile.ZipFile(file_path, 'r') as zip_file:
        csv_files = [f for f in zip_file.namelist() if f.endswith('.csv')]
        if len(csv_files) > 0:
            with zip_file.open(csv_files[0]) as csv_file:
                return pd.read_csv(csv_file)
        else:
            return None



def filter_by_contracts(df, contracts):
    df['CONTRACT'] = df['CONTRACT'].str.split(" ").str[1]
    return df[df['CONTRACT'].isin(contracts)]


def save_to_excel(df, file_path):
    if not file_path.endswith('.xlsx'):
        file_path += '.xlsx'
    df.to_excel(file_path, index=False)
    print(f"DataFrame saved as Excel file at {os.path.abspath(file_path)}")

def write_to_excel(data):
    
    #file_path = "V:\\Work Program\\P - Power BI\\Assets\\data_frame.xlsx"
    file_path =  "C:\\Users\\mdmitriev\\OneDrive - Osmose Utilities Services\\Documents\\Python Projects\\scripts\\Automation\\Western Power\\data_frame.xlsx"
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    
    try:
        # try to open an existing file
        book = load_workbook(file_path)
        writer.book = book
        data.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row+1)
        writer.save()
    except FileNotFoundError:
        # file does not exist yet, so create it
        data.to_excel(writer, index=False)
        writer.save()

def remove_duplicate_rows(data):
    return data.drop_duplicates()

def clean_pick_id(df):
    df['PICK_ID'] = df['PICK_ID'].str.split(" ").str[1]
    return df

def sort_through_dfs(contracts, list_of_data_frames, filtered_list):
    for data_frame in list_of_data_frames:
        if data_frame is not None:
                #contracts = ['C46607'] # as a hard-coded list of contract Nums
            df1 = filter_by_contracts(data_frame, contracts)
            df2 = clean_pick_id(df1)
            filtered_list.append(df2)
    return filtered_list

def sort_through_df(df, contracts, filtered_list):
    if df is not None:
        df = filter_by_contracts(df, contracts)
        df = clean_pick_id(df)
        filtered_list.append(df)
    return filtered_list


def main():
    
    folder_path ='./'
    list_of_data_frames = []
    filtered_list = []
    contracts = []
    # Define the folder path
    folder_path = "C:\\Users\\mdmitriev\\OneDrive - Osmose Utilities Services\\Documents\\Python Projects\\scripts\\Automation\\Western Power\\Trial"

    #if len(sys.argv) < 2:
    #    print("Please specify the name of the zip file as an argument")
    #    return
    #file_name = sys.argv[1]

    contract_list = sys.argv[1:] # using the remaining arguments to specify the contracts
    for contract in contract_list:
        contracts.append(str(contract))
    contracts_str = ','.join(contract_list)
    print(contracts)


    # Traverse through every file and subfolder in the directory
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            print(file)
            # Construct the file path
            file_path = os.path.join(root, file)
            # Check if the file is a CSV file
            if file.endswith(".csv"):
                # Read the CSV file into a data frame
                df = pd.read_csv(file_path, encoding='unicode_escape', low_memory=False)
                # Append the data frame to the list of data frames
                #list_of_data_frames.append(df)
                #filtered_list = sort_through_df(df, contracts, filtered_list)
                #df = filter_by_contracts(df, contracts)
                df = clean_pick_id(df)
                filtered_list.append(df)
            # Check if the file is a ZIP file
            elif file.endswith(".zip"):
                with zipfile.ZipFile(file_path, 'r') as zip_file:
                    csv_files = [f for f in zip_file.namelist() if f.endswith('.csv')]
                    if len(csv_files) > 0:
                        with zip_file.open(csv_files[0]) as csv_file:
                            df= pd.read_csv(csv_file, encoding='unicode_escape', low_memory=False)
                            #list_of_data_frames.append(df)
                            #filtered_list = sort_through_df(df, contracts, filtered_list)
                            #df = filter_by_contracts(df, contracts)
                            df = clean_pick_id(df)
                            filtered_list.append(df)
    print(len(filtered_list))
    #filtered_data_frames = sort_through_dfs(contracts, list_of_data_frames, filtered_data_frames)
                



    result_df = pd.concat(filtered_list, ignore_index=False)        
    save_to_excel(result_df, f"V:\\Work Program\\P - Power BI\\Assets\\data_frame_{contracts_str}")


if __name__ == '__main__':
    main()